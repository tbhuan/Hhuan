import csv


def writecsv(path,data):
    with open(path, "w") as f:
        writer = csv.writer(f)
        for rowData in data:
            print("rowData=", rowData)
            writer.writerow(rowData)

path = r"E:\\Python\\py17\\automatictext\\000001.csv"
writecsv(path, [[1, 2, 3], [4, 5, 6], [7, 8, 9]])

def readcev(path):
    infolist = []
    with open(path, "r") as f:
        allFile = csv.reader(f)
        for row in allFile:
            infolist.append(row)
    return infolist


path = r"E:\\Python\\py17\\automatictext\\PCB3.csv"
info = readcev(path)

import win32com
import win32com.client


def readWordFile(path):
    # 调用系统word功能，可以处理doc和docx两种文件
    mw = win32com.client.Dispatch("Word.Application")
    # 打开文件
    doc = mw.Documents.Open(path)
    for paragraph in doc.Paragraphs:
        line = paragraph.Range.Text
        print(line)
    doc.Close()
    mw.Quit()


path = r"E:\\Python\\py17\\Keyboardtext\\001.docx"

import win32com
import win32com.client


def readWordFiletootherFile(path, topath):
    mw = win32com.client.Dispatch("Word.Application")
    doc = mw.Documents.Open(path)
    # 将word的数据保存在另一个文件
    doc.SaveAs(topath, 2)
    doc.Close()
    mw.Quit()


path = r"E:\\Python\\py17\\Keyboardtext\\001.docx"
topath = r"E:\\Python\\py17\\Keyboardtext\\a.txt"

import win32com
import win32com.client
import os


def makeWordFile(path, name):
    word = win32com.client.Dispatch("Word.Application")
    # 让文档可见
    word.Visible = True
    # 创建文档
    doc = word.Documents.Add()
    # 写内容从头开始写
    r = doc.Range(0, 0)
    r.InsertAfter("你好，" + name + "\n")
    r.InsertAfter("python\n")
    # 存储文件
    doc.SaveAs(path)
    doc.Close()
    word.Quit()


names = ["张三", "李四", "王五"]
for name in names:
    path = os.path.join(os.getcwd(), name)
    makeWordFile(path, name)
# 有序字典
from collections import OrderedDict
# 存储数据
from pyexcel_xls import save_data


def makeExcelFile(path, data):
    dic = OrderedDict()
    for sheetNum, sheetValue in data.items():
        d = {}
        d[sheetNum] = sheetValue
        dic.update(d)

    save_data(path, dic)


path = r"E:\\Python\\py17\\automatictext\\b.xlsx"
makeExcelFile(path, {"表1": [[1, 2, 3], [4, 5, 6], [7, 8, 9]],
                     "表2": [[11, 22, 33], [44, 55, 66],
                            [77, 88, 99]]})

from openpyxl.reader.excel import load_workbook


def readXlsxFile(path):
    file = load_workbook(filename=path)
    print(file.get_sheet_names)
    sheets = file.get_sheet_names()
    sheet = file.get_sheet_by_name(sheets[0])
    for lineNum in range(1, sheet.max_row + 1):
        lineList = []
        print(sheet.max_row, sheet.max_column)
        for columnNum in range(1, sheet.max_column + 1):
            # 拿数据
            value = sheet.cell(row=lineNum,
                               column=columnNum).value
            if value != None:
                lineList.append(value)
        print(lineList)


path = r"E:\\Python\\py17\\automatictext\\001.xlsx"
readXlsxFile(path)

from openpyxl.reader.excel import load_workbook


def readXlsxFile(path):
    dic = {}
    file = load_workbook(filename=path)
    sheets = file.get_sheet_names()
    print(len(sheets))
    for sheetName in sheets:
        sheet = file.get_sheet_by_name(sheetName)
        # 一张表的所有数据
        sheetInfo = []
        for lineNum in range(1, sheet.max_row + 1):
            lineList = []
            for columnNum in range(1, sheet.max_column + 1):
                value = sheet.cell(row=lineNum,
                                   column=columnNum).value
                lineList.append(value)
            sheetInfo.append(lineList)
            # 将一张表的数据存到字典
            dic[sheetName] = sheetInfo
        return dic


path = r"E:\\Python\\py17\\automatictext\\001.xlsx"
dic = readXlsxFile(path)
print(dic)

# 有序字典
from collections import OrderedDict
# 读取数据
from pyexcel_xls import get_data


def readXlsAndXlsxFile(path):
    dic = OrderedDict()
    # 抓取数据
    xdata = get_data(path)
    for sheet in xdata:
        dic[sheet] = xdata[sheet]
    return dic


path = r"E:\\Python\\py17\\automatictext\\001.xlsx"
dic = readXlsAndXlsxFile(path)
print(dic)
print(len(dic))

import win32com
import win32com.client


def makeppt(path):
    ppt = win32com.client.Dispatch("PowerPoint.Application")
    ppt.Visible = True
    pptFile = ppt.Presentations.Add()

    # 创建页
    page1 = pptFile.Slides.Add(1, 1)
    t1 = page1.Shapes[0].TextFrame.TextRange
    t1.Text = "sunck"
    t2 = page1.Shapes[1].TextFrame.TextRange
    t2.Text = "sunck is a good man"
    # 保存
    pptFile.SaveAs(path)
    pptFile.Close()
    ppt.Quit()


path = r"E:\\Python\\py17\\automatictext\\sunk.ppt"
makeppt(path)

import win32com
import win32com.client


def makeppt(path):
    ppt = win32com.client.Dispatch("PowerPoint.Application")
    ppt.Visible = True
    pptFile = ppt.Presentations.Add()

    # 创建页
    page1 = pptFile.Slides.Add(1, 1)
    t1 = page1.Shapes[0].TextFrame.TextRange
    t1.Text = "sunck"
    t2 = page1.Shapes[1].TextFrame.TextRange
    t2.Text = "sunck is a good man"
    # 保存
    pptFile.SaveAs(path)
    pptFile.Close()
    ppt.Quit()


path = r"E:\\Python\\py17\\automatictext\\sunk.ppt"
makeppt(path)
